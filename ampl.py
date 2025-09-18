import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import streamlit as st
import tempfile

st.set_page_config(page_title="Excel Combiner Tool", page_icon="📊", layout="centered")
st.title("📊 Excel Combiner Tool")

# Upload multiple Excel files
uploaded_files = st.file_uploader("Upload Excel files", type="xlsx", accept_multiple_files=True)

if uploaded_files:
    dates = []

    # Step 1: Extract dates from B5
    for uploaded_file in uploaded_files:
        try:
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
            date_value = ws["B5"].value
            wb.close()

            if date_value:
                parsed_date = pd.to_datetime(str(date_value), errors="coerce", dayfirst=True)
                if pd.notnull(parsed_date):
                    dates.append(parsed_date)
        except Exception as e:
            st.warning(f"⚠️ Could not read B5 from {uploaded_file.name}: {e}")

    # Step 2: Pick latest date
    if dates:
        max_date = max(dates)
        date_str = max_date.strftime("%d-%m-%Y")
    else:
        date_str = datetime.today().strftime("%d-%m-%Y")

    output_file_name = f"Combined-{date_str}.xlsx"

    # Use temporary directory for safe file handling
    temp_dir = tempfile.mkdtemp()
    output_path = os.path.join(temp_dir, output_file_name)

    # Step 3: Combine into one Excel file
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for uploaded_file in uploaded_files:
            try:
                xls = pd.ExcelFile(uploaded_file)
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(uploaded_file, sheet_name=sheet_name)

                    # Excel sheet name limit (31 chars)
                    final_sheet_name = sheet_name
                    if final_sheet_name in writer.sheets:
                        final_sheet_name = f"{sheet_name}_{uploaded_file.name}".replace(".xlsx", "")[:31]

                    df.to_excel(writer, sheet_name=final_sheet_name, index=False)

                st.write(f"✅ Added **{uploaded_file.name}**")
            except Exception as e:
                st.warning(f"⚠️ Skipping {uploaded_file.name} due to error: {e}")

    st.success(f"🎉 Files combined successfully into `{output_file_name}`")

    # Step 4: Download button
    with open(output_path, "rb") as f:
        st.download_button("⬇️ Download Combined File", f, file_name=output_file_name)
